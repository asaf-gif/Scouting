"""
graph/migrate.py — Migrate Excel prototype data into Neo4j

Reads:  data/Business Model mapping - V3 6.xlsx
Writes: 27 BusinessModel nodes
        26 Scalar nodes
        3  Technology nodes
        702 TransformationVector nodes
        702 HAS_TRANSITION relationships (BM → BM)
        1404 FROM_BIM / TO_BIM relationships (Vector → BM)
        ~78 IMPACTS relationships (Tech → Scalar, where impact ≠ 0)

Usage:
    python graph/migrate.py
    python graph/migrate.py --source "data/Business Model mapping - V3 6.xlsx"
    python graph/migrate.py --clear    # wipe existing migrated nodes first
"""

import os
import sys
import argparse
from datetime import datetime, timezone

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from dotenv import load_dotenv
from neo4j import GraphDatabase
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn

from graph.migrate_utils import (
    IMPACT_MAP, IMPACT_LABEL, parse_impact,
    is_scalar_id, is_section_header,
    bim_id_from_index, scalar_id_from_code, tech_id_from_index,
)

load_dotenv()
console = Console()

PROVENANCE = {
    "source": "excel_prototype",
    "created_by": "migration",
    "confidence": 1.0,
}

DEFAULT_SOURCE = "data/Business Model mapping - V3 6.xlsx"


# ─────────────────────────────────────────────
# READERS
# ─────────────────────────────────────────────

def read_business_models(wb) -> list[dict]:
    """Read 27 BusinessModel nodes from the Example Matrix header row."""
    ws = wb["Example Matrix"]
    rows = list(ws.iter_rows(values_only=True))

    # Row 0 is the header: first cell is the "FROM \\ TO" label, rest are BM names
    bm_names = [str(v).strip() for v in rows[0][1:] if v]

    now = datetime.now(timezone.utc).isoformat()
    return [
        {
            "bim_id": bim_id_from_index(i),
            "name": name,
            "status": "Active",
            "typical_margins": "Variable",
            "version": 1,
            "created_at": now,
            "updated_at": now,
            **PROVENANCE,
        }
        for i, name in enumerate(bm_names, 1)
    ]


def read_scalars(wb) -> list[dict]:
    """Read Scalar nodes from the Scalar View sheet."""
    ws = wb["Scalar View"]
    rows = list(ws.iter_rows(values_only=True))

    scalars = []
    current_group = "General"
    now = datetime.now(timezone.utc).isoformat()

    for row in rows:
        if not any(row):
            continue

        cell0 = row[0]

        # Section header (e.g. "DEMAND — Customer-Side Signals")
        if is_section_header(cell0):
            current_group = str(cell0).strip()
            continue

        # Scalar row (e.g. A1, B3, E7)
        if is_scalar_id(cell0):
            code = str(cell0).strip().upper()
            name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            rationale = str(row[2]).strip() if len(row) > 2 and row[2] else ""

            scalars.append({
                "scalar_id": scalar_id_from_code(code),
                "code": code,
                "name": name,
                "description": name,
                "rationale": rationale,
                "group": current_group,
                "status": "Active",
                "version": 1,
                "created_at": now,
                "updated_at": now,
                **PROVENANCE,
            })

    return scalars


def read_technologies(wb) -> tuple[list[dict], list[dict]]:
    """
    Read 3 Technology nodes and their scalar IMPACTS from the Technology Impact sheet.
    Returns (technologies, impacts).
    """
    now = datetime.now(timezone.utc).isoformat()

    technologies = [
        {
            "tech_id": "TECH_001",
            "name": "GNNs — Graph Neural Networks",
            "short_name": "GNNs",
            "description": (
                "Relational pattern learning on graph-structured data. "
                "Amplifies network effects and relationship-dense business models."
            ),
            "category": "AI/ML",
            "tracking_status": "Active",
            "maturity_level": 65.0,
            "maturity_source": "Papers with Code — GNN benchmarks",
            "version": 1,
            "created_at": now,
            "updated_at": now,
            **PROVENANCE,
        },
        {
            "tech_id": "TECH_002",
            "name": "KGGen — Knowledge Graph Generation",
            "short_name": "KGGen",
            "description": (
                "Automated extraction of structured facts from unstructured text. "
                "Converts project insights into permanent, queryable knowledge assets."
            ),
            "category": "AI/ML",
            "tracking_status": "Active",
            "maturity_level": 55.0,
            "maturity_source": "KGGen benchmark on MTEB dataset",
            "version": 1,
            "created_at": now,
            "updated_at": now,
            **PROVENANCE,
        },
        {
            "tech_id": "TECH_003",
            "name": "Synthetic Audiences",
            "short_name": "SynthAudiences",
            "description": (
                "AI-generated consumer persona simulations. "
                "Expands market access while commoditising traditional primary research."
            ),
            "category": "AI/ML",
            "tracking_status": "Active",
            "maturity_level": 50.0,
            "maturity_source": "Accuracy vs. real panel benchmarks",
            "version": 1,
            "created_at": now,
            "updated_at": now,
            **PROVENANCE,
        },
    ]

    # Technology Impact sheet: col 0 = scalar code, cols 2/3/4 = GNNs/KGGen/SynthAudiences
    ws = wb["Technology Impact"]
    rows = list(ws.iter_rows(values_only=True))

    impacts = []
    tech_ids = ["TECH_001", "TECH_002", "TECH_003"]

    for row in rows:
        if not row or not is_scalar_id(row[0]):
            continue

        code = str(row[0]).strip().upper()
        scalar_id = scalar_id_from_code(code)

        for col_offset, tech_id in enumerate(tech_ids):
            col_idx = 2 + col_offset
            raw = row[col_idx] if col_idx < len(row) else None
            score, label = parse_impact(raw)

            impacts.append({
                "tech_id": tech_id,
                "scalar_id": scalar_id,
                "impact_score": score,
                "impact_level": label,
                "raw_value": str(raw).strip() if raw is not None else "",
                "created_at": now,
                **PROVENANCE,
            })

    return technologies, impacts


def read_vectors(wb, bms: list[dict]) -> list[dict]:
    """Read 702 TransformationVector nodes from the Example Matrix cells."""
    ws = wb["Example Matrix"]
    rows = list(ws.iter_rows(values_only=True))

    # Build name → bim_id lookup (stripped)
    name_to_id = {bm["name"].strip(): bm["bim_id"] for bm in bms}

    to_names = [str(v).strip() for v in rows[0][1:] if v is not None]

    now = datetime.now(timezone.utc).isoformat()
    vectors = []

    for data_row in rows[1:]:
        if not data_row[0]:
            continue
        from_name = str(data_row[0]).strip()
        from_id = name_to_id.get(from_name)
        if not from_id:
            continue

        for col_idx, to_name in enumerate(to_names):
            to_id = name_to_id.get(to_name.strip())
            if not to_id or from_id == to_id:
                continue  # skip self-transitions

            raw_cell = data_row[col_idx + 1] if (col_idx + 1) < len(data_row) else None
            example_text = str(raw_cell).strip() if raw_cell else ""

            vectors.append({
                "vector_id": f"VEC_{from_id}_{to_id}",
                "from_bim": from_id,
                "to_bim": to_id,
                "example_text": example_text,
                "evidence_quality": "Moderate" if example_text else "Speculative",
                "composite_score": 0.0,   # populated by scorer in Part 14
                "version": 1,
                "created_at": now,
                "updated_at": now,
                **PROVENANCE,
            })

    return vectors


# ─────────────────────────────────────────────
# TECH SCORES FROM TOP TRANSITIONS SHEET
# ─────────────────────────────────────────────

def _parse_score(raw) -> int:
    """Convert score strings like '+15', '○', '-2' to integers."""
    if raw is None:
        return 0
    s = str(raw).strip()
    if s in ("○", "—", "", "None"):
        return 0
    try:
        return int(s.replace("+", ""))
    except ValueError:
        return 0


def read_tech_scores(wb, bms: list[dict]) -> dict:
    """
    Read pre-computed transition scores from the 'Top Transitions' sheet.
    Returns: {vector_id: {tech_score_gnns, tech_score_kggen, tech_score_synthetic}}
    Only top/bottom ranked rows are present; all others default to 0.
    """
    name_to_id = {bm["name"].strip(): bm["bim_id"] for bm in bms}
    ws = wb["Top Transitions"]
    rows = list(ws.iter_rows(values_only=True))

    scores: dict[str, dict] = {}
    tech_key = None  # current technology field name

    for row in rows:
        if not any(row):
            continue

        cell0 = row[0]

        # Detect technology section headers
        if isinstance(cell0, str):
            if "TECHNOLOGY 1" in cell0 or "GNNs" in cell0 and "TECHNOLOGY" in cell0:
                tech_key = "tech_score_gnns"
                continue
            if "TECHNOLOGY 2" in cell0 or "KGGen" in cell0 and "TECHNOLOGY" in cell0:
                tech_key = "tech_score_kggen"
                continue
            if "TECHNOLOGY 3" in cell0 or "Synthetic" in cell0 and "TECHNOLOGY" in cell0:
                tech_key = "tech_score_synthetic"
                continue

        # Data rows: rank is an integer in col 0
        if tech_key and isinstance(cell0, int):
            from_name = str(row[1]).strip() if row[1] else None
            to_name   = str(row[2]).strip() if row[2] else None
            score_raw = row[3]

            if not from_name or not to_name:
                continue

            from_id = name_to_id.get(from_name)
            to_id   = name_to_id.get(to_name)
            if not from_id or not to_id or from_id == to_id:
                continue

            vid = f"VEC_{from_id}_{to_id}"
            if vid not in scores:
                scores[vid] = {"tech_score_gnns": 0, "tech_score_kggen": 0, "tech_score_synthetic": 0}
            scores[vid][tech_key] = _parse_score(score_raw)

    return scores


def write_tech_scores(driver, scores: dict):
    """Patch tech scores onto existing TransformationVector nodes."""
    with driver.session() as s:
        for vid, sc in scores.items():
            s.run("""
                MATCH (v:TransformationVector {vector_id: $vid})
                SET v.tech_score_gnns      = $gnns,
                    v.tech_score_kggen     = $kggen,
                    v.tech_score_synthetic = $synth
            """, vid=vid,
                 gnns=sc["tech_score_gnns"],
                 kggen=sc["tech_score_kggen"],
                 synth=sc["tech_score_synthetic"])
    console.print(f"  [green]✓[/green] Tech scores patched on {len(scores)} TransformationVector nodes")


# ─────────────────────────────────────────────
# WRITERS
# ─────────────────────────────────────────────

def get_driver():
    return GraphDatabase.driver(
        os.getenv("NEO4J_URI"),
        auth=(os.getenv("NEO4J_USER", "neo4j"), os.getenv("NEO4J_PASSWORD")),
    )


def clear_migrated(driver):
    console.print("[yellow]Clearing existing excel_prototype nodes...[/yellow]")
    with driver.session() as session:
        for label in ["BusinessModel", "Scalar", "TransformationVector", "Technology"]:
            r = session.run(
                f"MATCH (n:{label} {{source: 'excel_prototype'}}) DETACH DELETE n RETURN count(n) AS n"
            )
            console.print(f"  Deleted {r.single()['n']} {label} nodes")


def write_all(driver, bms, scalars, technologies, impacts, vectors):
    now = datetime.now(timezone.utc).isoformat()

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        console=console,
    ) as progress:

        # 1. BusinessModel nodes
        t = progress.add_task("BusinessModel nodes", total=len(bms))
        with driver.session() as s:
            for bm in bms:
                s.run("MERGE (n:BusinessModel {bim_id:$id}) SET n += $p",
                      id=bm["bim_id"], p=bm)
                progress.advance(t)
        console.print(f"  [green]✓[/green] {len(bms)} BusinessModel nodes")

        # 2. Scalar nodes
        t = progress.add_task("Scalar nodes", total=len(scalars))
        with driver.session() as s:
            for sc in scalars:
                s.run("MERGE (n:Scalar {scalar_id:$id}) SET n += $p",
                      id=sc["scalar_id"], p=sc)
                progress.advance(t)
        console.print(f"  [green]✓[/green] {len(scalars)} Scalar nodes")

        # 3. Technology nodes
        t = progress.add_task("Technology nodes", total=len(technologies))
        with driver.session() as s:
            for tech in technologies:
                s.run("MERGE (n:Technology {tech_id:$id}) SET n += $p",
                      id=tech["tech_id"], p=tech)
                progress.advance(t)
        console.print(f"  [green]✓[/green] {len(technologies)} Technology nodes")

        # 4. IMPACTS relationships
        t = progress.add_task("IMPACTS relationships", total=len(impacts))
        with driver.session() as s:
            for imp in impacts:
                s.run("""
                    MATCH (t:Technology {tech_id:$tid})
                    MATCH (sc:Scalar {scalar_id:$sid})
                    MERGE (t)-[r:IMPACTS]->(sc)
                    SET r.impact_score  = $score,
                        r.impact_level  = $level,
                        r.raw_value     = $raw,
                        r.source        = 'excel_prototype',
                        r.created_at    = $ts
                """, tid=imp["tech_id"], sid=imp["scalar_id"],
                     score=imp["impact_score"], level=imp["impact_level"],
                     raw=imp["raw_value"], ts=now)
                progress.advance(t)
        console.print(f"  [green]✓[/green] {len(impacts)} IMPACTS relationships")

        # 5. TransformationVector nodes + HAS_TRANSITION + FROM_BIM/TO_BIM
        t = progress.add_task("TransformationVector nodes", total=len(vectors))
        with driver.session() as s:
            for vec in vectors:
                s.run("""
                    MERGE (v:TransformationVector {vector_id:$vid})
                    SET v += $p
                    WITH v
                    MATCH (f:BusinessModel {bim_id:$fid})
                    MATCH (to:BusinessModel {bim_id:$tid})
                    MERGE (f)-[:HAS_TRANSITION]->(to)
                    MERGE (v)-[:FROM_BIM]->(f)
                    MERGE (v)-[:TO_BIM]->(to)
                """, vid=vec["vector_id"], p=vec,
                     fid=vec["from_bim"], tid=vec["to_bim"])
                progress.advance(t)
        console.print(f"  [green]✓[/green] {len(vectors)} TransformationVector nodes")
        console.print(f"  [green]✓[/green] {len(vectors)} HAS_TRANSITION relationships")
        console.print(f"  [green]✓[/green] {len(vectors) * 2} FROM_BIM/TO_BIM relationships")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Migrate Excel prototype to Neo4j")
    parser.add_argument("--source", default=DEFAULT_SOURCE)
    parser.add_argument("--clear", action="store_true",
                        help="Delete existing excel_prototype nodes before migrating")
    args = parser.parse_args()

    console.print(f"\n[bold]Systematic Problem Scouting — Data Migration[/bold]")
    console.print(f"Source: [cyan]{args.source}[/cyan]\n")

    wb = openpyxl.load_workbook(args.source, data_only=True)

    console.print("Reading Excel...")
    bms = read_business_models(wb)
    scalars = read_scalars(wb)
    technologies, impacts = read_technologies(wb)
    vectors = read_vectors(wb, bms)

    console.print(f"  {len(bms)} business models")
    console.print(f"  {len(scalars)} scalars")
    console.print(f"  {len(technologies)} technologies · {len(impacts)} impact mappings")
    console.print(f"  {len(vectors)} transition vectors\n")

    driver = get_driver()

    if args.clear:
        clear_migrated(driver)
        console.print()

    console.print("Reading tech scores from Top Transitions sheet...")
    tech_scores = read_tech_scores(wb, bms)
    console.print(f"  {len(tech_scores)} vectors with explicit tech scores\n")

    console.print("[bold]Writing to Neo4j...[/bold]")
    write_all(driver, bms, scalars, technologies, impacts, vectors)

    console.print("Patching tech scores...")
    write_tech_scores(driver, tech_scores)

    driver.close()
    console.print("\n[bold green]Migration complete.[/bold green]")
    console.print("Validate: [cyan]python tests/test_03_migration.py[/cyan]")


if __name__ == "__main__":
    main()
